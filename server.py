#!/usr/bin/env python3
"""
KZ Business Database - Local Server
- Читает все .xlsx из LOCAL_DB_PATH при старте
- Сохраняет ВСЕ изменения (добавить/удалить/редактировать) обратно в файл database.xlsx
- Эндпоинты: GET /api/load-local, POST /api/save-all
"""

import os, json, glob
from http.server import HTTPServer, SimpleHTTPRequestHandler
import google.generativeai as genai

# ===================== КОНФИГУРАЦИЯ =====================
API_KEY       = os.environ.get("GEMINI_API_KEY", "")
LOCAL_DB_PATH = r"C:\Users\galim\OneDrive\Desktop\kz-db\database"
DB_FILE       = os.path.join(LOCAL_DB_PATH, "database.xlsx")   # единый файл для записи

COLUMNS = [
    'Company name', 'Category', 'Status', 'City', 'Website', 'Email',
    'Phone', 'Address', 'CEO-1', 'Position-1', 'CEO-2', 'Position-2',
    'Linkedin', 'Status-L', 'Facebook', 'Status-F'
]

model = None


# ===================== EXCEL — ЧТЕНИЕ =====================
def load_local_xlsx():
    """Читает все .xlsx из LOCAL_DB_PATH, объединяет, возвращает список dict."""
    try:
        import openpyxl
    except ImportError:
        print("  ⚠️  pip install openpyxl")
        return []

    if not os.path.isdir(LOCAL_DB_PATH):
        print(f"  ⚠️  Папка не найдена: {LOCAL_DB_PATH}")
        return []

    files = glob.glob(os.path.join(LOCAL_DB_PATH, "*.xlsx"))
    if not files:
        print(f"  ⚠️  .xlsx не найдены в: {LOCAL_DB_PATH}")
        return []

    COL_MAP = {
        'Company name': ['Company name', 'Название компании', 'название компании'],
        'Category':     ['Category', 'Category ', 'Категория'],
        'Status':       ['Status', 'Status ', 'Статус'],
        'Website':      ['Website', 'Сайт'],
        'Email':        ['Email'],
        'City':         ['City', 'Город'],
        'Phone':        ['Phone', 'Phone, contacts', 'Phone, contacts ', 'Телефон'],
        'Address':      ['Address', 'Адрес'],
        'CEO-1':        ['CEO-1', 'Председатель правления / CEO'],
        'Position-1':   ['Position-1', 'Должность.1'],
        'CEO-2':        ['CEO-2', 'Глава совета директоров'],
        'Position-2':   ['Position-2', 'Должность'],
        'Linkedin':     ['Linkedin', 'LinkedIn'],
        'Status-L':     ['Status-L', 'Статус LinkedIn'],
        'Facebook':     ['Facebook'],
        'Status-F':     ['Status-F', 'Статус Facebook'],
    }

    all_records = []
    seen_names  = set()

    for filepath in sorted(files):
        fname = os.path.basename(filepath)
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

            for row in ws.iter_rows(min_row=2, values_only=True):
                raw = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row) if i < len(headers)}
                rec = {}
                for target, aliases in COL_MAP.items():
                    val = ""
                    for alias in aliases:
                        if raw.get(alias):
                            val = raw[alias]; break
                    rec[target] = val

                name = rec.get('Company name', '').strip()
                if name and name.lower() not in seen_names:
                    seen_names.add(name.lower())
                    all_records.append(rec)

            wb.close()
            print(f"  ✓ {fname}: загружено записей")
        except Exception as e:
            print(f"  ⚠️  Ошибка {fname}: {e}")

    print(f"  ✓ Итого: {len(all_records)} компаний")
    return all_records


# ===================== EXCEL — ЗАПИСЬ =====================
def save_all_xlsx(records: list):
    """Перезаписывает database.xlsx полным списком компаний."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  ⚠️  pip install openpyxl")
        return False

    os.makedirs(LOCAL_DB_PATH, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Database"

    # Заголовки
    header_fill = PatternFill("solid", fgColor="C8622A")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Данные
    for rec in records:
        ws.append([str(rec.get(col, "")) for col in COLUMNS])

    # Ширина колонок
    col_widths = {
        'Company name': 35, 'Category': 25, 'Status': 12, 'City': 15,
        'Website': 30, 'Email': 28, 'Phone': 18, 'Address': 35,
        'CEO-1': 25, 'Position-1': 30, 'CEO-2': 25, 'Position-2': 30,
        'Linkedin': 22, 'Status-L': 14, 'Facebook': 22, 'Status-F': 14,
    }
    for i, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = col_widths.get(col, 18)

    wb.save(DB_FILE)
    print(f"  ✓ Сохранено {len(records)} записей → {DB_FILE}")
    return True


# ===================== GEMINI =====================
def get_model():
    global model
    if not model:
        if not API_KEY:
            raise ValueError("GEMINI_API_KEY не задан")
        genai.configure(api_key=API_KEY)
        model = genai.GenerativeModel(
            model_name="gemini-2.5-flash",
            generation_config={"temperature": 0.1, "response_mime_type": "application/json"}
        )
    return model


# ===================== HTTP HANDLER =====================
class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=os.path.dirname(os.path.abspath(__file__)), **kwargs)

    def log_message(self, format, *args):
        print(f"  {self.command} {self.path.split('?')[0]} → {args[1] if len(args) > 1 else ''}")

    def do_OPTIONS(self):
        self.send_response(200); self._cors(); self.end_headers()

    def do_GET(self):
        if self.path == "/api/load-local":
            self._handle_load()
        else:
            super().do_GET()

    def do_POST(self):
        routes = {
            "/api/save-all":  self._handle_save_all,
            "/api/search":    self._handle_search,
        }
        handler = routes.get(self.path)
        if handler:
            handler()
        else:
            self.send_error(404)

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS, GET")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def _json_response(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status); self._cors()
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers(); self.wfile.write(body)

    def _read_body(self):
        length = int(self.headers.get("Content-Length", 0))
        return json.loads(self.rfile.read(length))

    # GET /api/load-local
    def _handle_load(self):
        rows = load_local_xlsx()
        self._json_response({"success": True, "data": rows, "count": len(rows)})

    # POST /api/save-all  — фронт присылает весь массив data[]
    def _handle_save_all(self):
        try:
            body    = self._read_body()
            records = body.get("data", [])
            ok      = save_all_xlsx(records)
            self._json_response({"success": ok, "saved": len(records)})
        except Exception as e:
            print(f"  ⚠️  save-all: {e}")
            self._json_response({"error": str(e)}, 500)

    # POST /api/search  — Gemini ищет компанию
    def _handle_search(self):
        try:
            body         = self._read_body()
            company_name = body.get("company", "").strip()
            categories   = body.get("categories", "")
            if not company_name:
                self._json_response({"error": "Введите название"}, 400); return

            m      = get_model()
            prompt = f"""You are a professional business researcher.
Return a JSON object about the Kazakhstan company: "{company_name}".
Pick Category from: {categories}.
Use exact keys: "Company name","Category","City","Website","Email","Phone","Address",
"CEO-1","Position-1","CEO-2","Position-2","Linkedin","Facebook","Status-L","Status-F".
For social media use only handles. If not found use "". Return ONLY JSON."""

            resp   = m.generate_content(prompt)
            parsed = json.loads(resp.text.strip())
            parsed["Status"] = "Активный"
            if not parsed.get("Company name"):
                parsed["Company name"] = company_name

            self._json_response({"success": True, "data": parsed})
        except Exception as e:
            print(f"  ⚠️  search: {e}")
            self._json_response({"error": str(e)}, 500)


# ===================== UTILS =====================
def load_env():
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(path):
        with open(path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ[k.strip()] = v.strip().strip('"').strip("'")
        print("  ✓ Загружен .env")


# ===================== MAIN =====================
def main():
    load_env()
    global API_KEY
    API_KEY = os.environ.get("GEMINI_API_KEY", API_KEY)

    print("\n" + "=" * 60)
    print("  🗄  KZ Business Database — Local XLSX Server")
    print("=" * 60)

    if not API_KEY:
        print("  ⚠️  GEMINI_API_KEY не задан!")

    print(f"\n  📂 База: {LOCAL_DB_PATH}")
    rows = load_local_xlsx()
    print(f"  ✓ Загружено: {len(rows)} компаний\n")

    import socket
    try:
        local_ip = socket.gethostbyname(socket.gethostname())
    except Exception:
        local_ip = "ВАШ_IP"

    port   = 8000
    server = HTTPServer(("0.0.0.0", port), Handler)

    print(f"  ✓ Локально:        http://localhost:{port}")
    print(f"  ✓ По сети (Wi-Fi): http://{local_ip}:{port}")
    print(f"  ✓ Для интернета:   ngrok http {port}")
    print("\n  Ctrl+C для остановки\n")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Сервер остановлен.")


if __name__ == "__main__":
    main()
